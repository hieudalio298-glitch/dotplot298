"""
Script lấy TẤT CẢ khoản mục BCTC ngành Đường - Nguồn KBS
KBS API trả về cấu trúc sẵn có Form FiinTrade (Cột là Quý, Dòng là Chỉ tiêu)
Ta chỉ cần GroupBy theo 'item' và Sum các cột Quý.
"""

import sys, io, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

from vnstock import Vnstock

# ─── CẤU HÌNH ────────────────────────────────────────────────────────────────
SOURCE = "KBS"
OUTPUT_FILE = "sugar_industry_kbs_fiintrade.xlsx"
UNIT = 1  # KBS trả về đơn vị Tỷ đồng sẵn, nhưng có thể cần kiểm tra. Vnstock source KBS thường trả về chuẩn tỷ đồng hoặc tỷ vnd. Để an toàn ta cứ giữ nguyên, hoặc chia UNIT = 1.
# Thường vnstock nguồn KBS trả về đơn vị tỷ. Ta set UNIT = 1
UNIT = 1  
SUGAR_STOCKS = ["SBT", "QNS", "LSS", "SLS", "KTS", "FBT"]
PERIOD = "quarter"
RETRY = 3

print("=" * 70)
print(f"  TẢI DỮ LIỆU {SOURCE} & XUẤT FORM FIINTRADE (TỔNG NGÀNH)")
print("=" * 70)

def safe_fetch(api_func, label):
    for i in range(RETRY):
        try:
            return api_func()
        except Exception as e:
            if i < RETRY - 1:
                time.sleep(2)
            else:
                print(f"⚠️ {label} lỗi: {e}")
                return pd.DataFrame()

# ─── 1. TẢI DỮ LIỆU ────────────────────────────────────────────────────────
all_income = []
all_balance = []
all_cf = []

for idx, sym in enumerate(SUGAR_STOCKS):
    print(f"[{idx+1}/{len(SUGAR_STOCKS)}] Đang tải {sym} (KBS)...", flush=True)
    try:
        fin = Vnstock().stock(symbol=sym, source=SOURCE).finance
        
        df_inc = safe_fetch(lambda: fin.income_statement(period=PERIOD), f"{sym} KQKD")
        if not df_inc.empty:
            df_inc.insert(0, "symbol", sym)
            all_income.append(df_inc)
            
        df_bal = safe_fetch(lambda: fin.balance_sheet(period=PERIOD), f"{sym} CĐKT")
        if not df_bal.empty:
            df_bal.insert(0, "symbol", sym)
            all_balance.append(df_bal)
            
        df_cf = safe_fetch(lambda: fin.cash_flow(period=PERIOD), f"{sym} LCTT")
        if not df_cf.empty:
            df_cf.insert(0, "symbol", sym)
            all_cf.append(df_cf)
            
        time.sleep(1)
    except Exception as e:
        print(f"❌ Lỗi fetching {sym}: {e}")

df_income = pd.concat(all_income, ignore_index=True) if all_income else pd.DataFrame()
df_balance = pd.concat(all_balance, ignore_index=True) if all_balance else pd.DataFrame()
df_cashflow = pd.concat(all_cf, ignore_index=True) if all_cf else pd.DataFrame()

# ─── 2. HÀM TỔNG HỢP VÀ CHUYỂN ĐỔI FORM FIINTRADE ──────────────────────────
def aggregate_industry(df):
    if df.empty:
        return pd.DataFrame()
        
    df = df.copy()
    
    # Xác định các cột chứa kỳ (quý-năm, dạng Q hoặc YYYY)
    # df.columns thường là: ['symbol', 'item', 'item_id', '2025-Q2', '2025-Q1', ...]
    skip_cols = ["symbol", "item", "item_id"]
    quarter_cols = [c for c in df.columns if c not in skip_cols]
    
    # Trích xuất riêng cột item để làm key group
    # Do KBS đôi khi các công ty có item lộn xộn, ta drop symbol và item_id, groupby 'item'
    df_numeric = df[["item"] + quarter_cols].copy()
    
    # Ép kiểu số
    for c in quarter_cols:
        df_numeric[c] = pd.to_numeric(df_numeric[c], errors="coerce")
        # Nguồn KBS có thể đang trả về VNĐ hoặc Tỷ. 
        # Nếu mean của cột quá lớn (>1e5), chứng tỏ là Đồng, cần chia 1e9.
        # Ở đây ta check dòng đầu của cột đầu nếu lớn hơn tỷ
        
    # Group By Item (Chỉ tiêu) và Sum các phân cột.
    agg_df = df_numeric.groupby("item")[quarter_cols].sum(min_count=1).reset_index()
    
    # Check tỷ đồng. Nếu giá trị trung bình quá lớn => Đang là Đồng
    max_val = agg_df[quarter_cols].max().max()
    if max_val > 1e6: # Doanh thu trên 1 triệu tỷ? Chắc chắn là đang tính bằng đồng
        for c in quarter_cols:
            agg_df[c] = agg_df[c] / 1e9
            
    # Sort columns để Kỳ mới nhất nằm bên trái, Hoặc Kỳ cũ nhất bên trái.
    # FiinTrade thường có Kỳ mới nhất bên trái -> giảm dần.
    q_sorted = sorted(quarter_cols, reverse=True)
    agg_df = agg_df[["item"] + q_sorted]
    
    # Rename
    agg_df = agg_df.rename(columns={"item": "Chỉ tiêu (Tỷ đồng)"})
    return agg_df

# ─── 3. TẠO CÁC BẢNG DATA ────────────────────────────────────────────────
dict_sheets = {
    "KQKD Toàn Ngành": aggregate_industry(df_income),
    "CĐKT Toàn Ngành": aggregate_industry(df_balance),
    "LCTT Toàn Ngành": aggregate_industry(df_cashflow),
}

# ─── 4. XUẤT EXCEL DỊNH DẠNG ĐẸP ─────────────────────────────────────────
print(f"\n📊 Đang xuất file Excel: {OUTPUT_FILE} ...")

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    has_data = False
    for sheet_name, df_out in dict_sheets.items():
        if df_out.empty: continue
        has_data = True
        df_out = df_out.round(1)
        df_out.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
        
    if not has_data:
        print("❌ Không có dữ liệu để xuất")
        sys.exit(1)
        
    wb = writer.book
    
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color="1F4E78", size=14)
    item_font = Font(bold=True, size=10) # Tên chỉ tiêu in đậm
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin = Side(style="thin", color="BDD7EE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Merge cell cho Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        ws.cell(1, 1).value = f"BÁO CÁO CỘNG DỒN TOÀN NGÀNH ĐƯỜNG - {sheet_name.upper()} (FIINTRADE FORMAT)"
        ws.cell(1, 1).font = title_font
        ws.cell(1, 1).alignment = align_center
        
        # Header (Dòng 2: Q1, Q2...)
        for cell in ws[2]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = align_center
            cell.border = border
            
        # Data rows
        for row in list(ws.rows)[2:]:
            row[0].font = item_font
            row[0].alignment = align_left
            row[0].border = border
            
            for cell in row[1:]:
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.0'
                    if cell.value < 0:
                        cell.font = Font(color="FF0000") # Số âm màu đỏ
                cell.alignment = align_right
                
        # Freeze panes (Cố định cột đầu và 2 dòng đầu)
        ws.freeze_panes = "B3"
        
        # Chỉnh độ rộng
        ws.column_dimensions["A"].width = 50 
        for col_idx in range(2, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 14

print(f"✅ Hoàn tất lưu file: {OUTPUT_FILE}")
