"""
Script lấy TẤT CẢ khoản mục BCTC ngành Đường - Nguồn KBS
Format xuất Excel: Chuẩn FiinTrade (Dòng: Chỉ tiêu, Cột: Quý)
"""

import sys, io, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

from vnstock import Vnstock

# ─── CẤU HÌNH ────────────────────────────────────────────────────────────────
SOURCE = "KBS"
OUTPUT_FILE = "sugar_industry_kbs_fiintrade.xlsx"
UNIT = 1e9  # Tỷ đồng
SUGAR_STOCKS = ["SBT", "QNS", "LSS", "SLS", "KTS", "FBT"]
PERIOD = "quarter"
RETRY = 3

print("=" * 70)
print(f"  TẢI DỮ LIỆU {SOURCE} & XUẤT FORM FIINTRADE")
print("=" * 70)

def safe_fetch(api_func, label):
    for i in range(RETRY):
        try:
            return api_func()
        except Exception as e:
            if i < RETRY - 1:
                time.sleep(2)
            else:
                print(f"⚠️ {label} lỗi: {e}")
                return pd.DataFrame()

# ─── 1. TẢI DỮ LIỆU ────────────────────────────────────────────────────────
all_income = []
all_balance = []
all_cf = []

for idx, sym in enumerate(SUGAR_STOCKS):
    print(f"[{idx+1}/{len(SUGAR_STOCKS)}] Đang tải {sym} (KBS)...", flush=True)
    try:
        fin = Vnstock().stock(symbol=sym, source=SOURCE).finance
        
        # KQKD
        df_inc = safe_fetch(lambda: fin.income_statement(period=PERIOD), f"{sym} KQKD")
        if not df_inc.empty:
            df_inc.insert(0, "symbol", sym)
            all_income.append(df_inc)
            
        # CĐKT
        df_bal = safe_fetch(lambda: fin.balance_sheet(period=PERIOD), f"{sym} CĐKT")
        if not df_bal.empty:
            df_bal.insert(0, "symbol", sym)
            all_balance.append(df_bal)
            
        # LCTT
        df_cf = safe_fetch(lambda: fin.cash_flow(period=PERIOD), f"{sym} LCTT")
        if not df_cf.empty:
            df_cf.insert(0, "symbol", sym)
            all_cf.append(df_cf)
            
        time.sleep(1) # Tránh rate limit
    except Exception as e:
        print(f"❌ Không khởi tạo được API cho {sym}: {e}")

df_income = pd.concat(all_income, ignore_index=True) if all_income else pd.DataFrame()
df_balance = pd.concat(all_balance, ignore_index=True) if all_balance else pd.DataFrame()
df_cashflow = pd.concat(all_cf, ignore_index=True) if all_cf else pd.DataFrame()

print(f"\n✅ Đã tải xong: KQKD ({len(df_income)} dòng) | CĐKT ({len(df_balance)} dòng) | LCTT ({len(df_cashflow)} dòng)")

# ─── 2. HÀM CHUYỂN ĐỔI SANG FORM FIINTRADE ────────────────────────────────
def make_fiintrade_format(df, aggregate=True):
    if df.empty:
        return pd.DataFrame()
        
    df = df.copy()
    # Tạo label Quý
    df["Kỳ báo cáo"] = df["Năm"].astype(str) + "-Q" + df["Kỳ"].astype(str)
    
    skip_cols = ["symbol", "CP", "Năm", "Kỳ", "Kỳ báo cáo"]
    numeric_cols = []
    
    # Ép kiểu dữ liệu và lọc cột
    for c in df.columns:
        if c not in skip_cols:
            df[c] = pd.to_numeric(df[c], errors="coerce")
            if not df[c].isna().all():
                numeric_cols.append(c)
                df[c] = df[c] / UNIT  # Chuyển về Tỷ đồng
                
    if not numeric_cols:
        return pd.DataFrame()
        
    if aggregate:
        # Cộng dồn toàn ngành
        agg_df = df.groupby("Kỳ báo cáo")[numeric_cols].sum(min_count=1).reset_index()
    else:
        # Lấy theo công ty (không dùng trong script hiện tại, nhưng prep sẵn)
        agg_df = df.copy()
        
    # Form FiinTrade: Dòng là Chỉ tiêu, Cột là Kỳ báo cáo
    # Từ: Kỳ báo cáo | Doanh thu | Lợi nhuận ...
    # Sang: Chỉ tiêu | Q1 | Q2 | Q3 ...
    agg_df = agg_df.set_index("Kỳ báo cáo").T
    
    # Sắp xếp lại cột: Mới nhất nằm bên phải (hoặc bên trái tùy thói quen, mặc định FiinTrade thường sắp mới nhất bên phải hoặc trái. Chọn cũ -> mới từ trái sang phải cho chuẩn Time Series)
    cols = sorted(agg_df.columns.tolist()) 
    # Nhưng FiinTrade thường xếp MỚI NHẤT ở bên TRÁI. Ta sẽ xếp giảm dần.
    cols = sorted(agg_df.columns.tolist(), reverse=True)
    agg_df = agg_df[cols]
    
    agg_df.index.name = "Chỉ tiêu (Tỷ đồng)"
    return agg_df.reset_index()

# ─── 3. TẠO CÁC BẢNG DATA ────────────────────────────────────────────────
dict_sheets = {
    "KQKD Toàn Ngành": make_fiintrade_format(df_income),
    "CĐKT Toàn Ngành": make_fiintrade_format(df_balance),
    "LCTT Toàn Ngành": make_fiintrade_format(df_cashflow),
}

# ─── 4. XUẤT EXCEL DỊNH DẠNG ĐẸP ─────────────────────────────────────────
print(f"📊 Đang xuất file Excel: {OUTPUT_FILE} ...")

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    for sheet_name, df_out in dict_sheets.items():
        if df_out.empty: continue
        
        df_out = df_out.round(1)
        df_out.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
        
    wb = writer.book
    
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color="1F4E78", size=14)
    item_font = Font(bold=True, size=10) # Tên chỉ tiêu in đậm
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin = Side(style="thin", color="BDD7EE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Merge cell cho Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        ws.cell(1, 1).value = f"BÁO CÁO CỘNG DỒN TOÀN NGÀNH ĐƯỜNG - {sheet_name.upper()} (FIINTRADE FORMAT)"
        ws.cell(1, 1).font = title_font
        ws.cell(1, 1).alignment = align_center
        
        # Header (Dòng 2: Q1, Q2...)
        for cell in ws[2]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = align_center
            cell.border = border
            
        # Data rows
        for row in list(ws.rows)[2:]:
            # Cột đầu tiên (Tên chỉ tiêu)
            row[0].font = item_font
            row[0].alignment = align_left
            row[0].border = border
            
            # Cột giá trị
            for cell in row[1:]:
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.0'
                    if cell.value < 0:
                        cell.font = Font(color="FF0000") # Số âm màu đỏ
                cell.alignment = align_right
                
        # Freeze panes (Cố định cột đầu và 2 dòng đầu)
        ws.freeze_panes = "B3"
        
        # Chỉnh độ rộng
        ws.column_dimensions["A"].width = 45 # Cột Chỉ tiêu rộng hơn
        for col_idx in range(2, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 14

print(f"✅ Hoàn tất lưu file: {OUTPUT_FILE}")
