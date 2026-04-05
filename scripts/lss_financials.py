"""
Script lấy BCTC của riêng mã LSS - Tất cả các năm
Sử dụng nguồn KBS (Chuẩn format VAS Tiếng Việt)
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from vnstock import Vnstock

SYMBOL = "LSS"
OUTPUT_FILE = "lss_financials_all_years_vas.xlsx"
PERIOD = "year"
SOURCE = "KBS"

print("=" * 70)
print(f"  TẢI BCTC {SYMBOL} - TẤT CẢ CÁC NĂM (Nguồn {SOURCE} - Chuẩn VAS)")
print("=" * 70)

fin = Vnstock().stock(symbol=SYMBOL, source=SOURCE).finance

df_inc = fin.income_statement(period=PERIOD)
df_bal = fin.balance_sheet(period=PERIOD)
df_cf = fin.cash_flow(period=PERIOD)

def format_dataframe(df):
    if df is None or df.empty:
        return pd.DataFrame()
        
    df = df.copy()
    
    # KBS trả về cột: 'item', 'item_id', '2024', '2023'...
    # Xoá cột 'item_id' nếu có
    if "item_id" in df.columns:
        df = df.drop(columns=["item_id"])
        
    # Đổi tên cột 'item' -> 'Chỉ tiêu'
    if "item" in df.columns:
        df = df.rename(columns={"item": "Chỉ tiêu (VND)"})
        
    # Xác định các năm và xếp Năm mới nhất bên trái
    year_cols = [c for c in df.columns if str(c).isdigit()]
    year_cols_sorted = sorted(year_cols, reverse=True)
    
    # Thứ tự cột: [Chỉ tiêu] + [Các Năm]
    final_cols = ["Chỉ tiêu (VND)"] + year_cols_sorted
    df = df[final_cols]
    
    return df

dict_sheets = {
    "KQ Kinh Doanh": format_dataframe(df_inc),
    "Cân Đối Kế Toán": format_dataframe(df_bal),
    "Lưu Chuyển Tiền Tệ": format_dataframe(df_cf),
}

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    for sheet_name, df_out in dict_sheets.items():
        if df_out.empty: continue
        df_out.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
        
    wb = writer.book
    
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color="1F4E78", size=14)
    item_font = Font(bold=True, size=10)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin = Side(style="thin", color="BDD7EE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        ws.cell(1, 1).value = f"BÁO CÁO TÀI CHÍNH LSS DÀI HẠN - {sheet_name.upper()} (CHUẨN VAS)"
        ws.cell(1, 1).font = title_font
        ws.cell(1, 1).alignment = align_center
        
        for cell in ws[2]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = align_center
            cell.border = border
            
        for row in list(ws.rows)[2:]:
            row[0].font = item_font
            row[0].alignment = align_left
            row[0].border = border
            
            for cell in row[1:]:
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0' 
                    if cell.value < 0:
                        cell.font = Font(color="FF0000")
                cell.alignment = align_right
                
        ws.freeze_panes = "B3"
        
        ws.column_dimensions["A"].width = 50 
        for col_idx in range(2, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 16

print(f"✅ Hoàn tất lưu file: {OUTPUT_FILE} bằng nguồn KBS Chuẩn VAS!")
