"""
Tải BCTC LSS - TẤT CẢ LỊCH SỬ MAX CỦA VCI
Nguồn: VCI (Có tỷ lệ lịch sử dài tới 13 năm và 52 quý)
Format: Chuẩn VAS, ngang kiểu FiinTrade
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

from vnstock import Vnstock

SYMBOL = "LSS"
SOURCE = "VCI"

print("=" * 70)
print(f"  TẢI BCTC {SYMBOL} - TOÀN BỘ LỊCH SỬ TỐI ĐA (13 Năm / 52 Quý)")
print("=" * 70)

fin = Vnstock().stock(symbol=SYMBOL, source=SOURCE).finance

def fetch_and_export(period, output_file):
    print(f"\n[+] Đang tải dữ liệu {period.upper()} từ {SOURCE}...")
    
    # Lang="vi" để VCI trả về tên column Tiếng Việt
    df_inc = fin.income_statement(period=period, lang="vi")
    df_bal = fin.balance_sheet(period=period, lang="vi")
    df_cf = fin.cash_flow(period=period, lang="vi")
    
    def format_dataframe(df):
        if df is None or df.empty:
            return pd.DataFrame()
            
        df = df.copy()
        
        # VCI trả về dạng Bảng (Mỗi dòng là 1 Quý/Năm, Các cột là Chỉ tiêu)
        # VD cột: CP, Năm, Kỳ, Doanh thu thuần...
        
        if period == "quarter" and "Kỳ" in df.columns and "Năm" in df.columns:
            # Tạo index dạng Q1-2024
            df["Kỳ báo cáo"] = df["Kỳ"].astype(str) + "-" + df["Năm"].astype(str)
            df = df.set_index("Kỳ báo cáo")
        elif "Năm" in df.columns:
            df["Kỳ báo cáo"] = df["Năm"].astype(str)
            df = df.set_index("Kỳ báo cáo")
        else:
            return df
            
        # Xoá các cột không phải là số (ngoại trừ index)
        drop_cols = ["CP", "Năm", "Kỳ"]
        for c in drop_cols:
            if c in df.columns:
                df = df.drop(columns=[c])
                
        # Fill/Ép kiểu số
        for c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
            
        # Nghịch đảo lại thành FiinTrade format (Dòng: Chỉ tiêu, Cột: Kỳ)
        df_transposed = df.T
        
        # Sắp xếp các cột thời gian từ Mới Nhất -> Cũ Nhất
        def sort_key(col_name):
            val = str(col_name)
            if "-" in val:
                q, y = val.split("-")
                return int(y) * 10 + int(q)
            return int(val)
            
        sorted_cols = sorted(df_transposed.columns, key=sort_key, reverse=True)
        df_transposed = df_transposed[sorted_cols]
        
        df_transposed.index.name = "Chỉ tiêu (VNĐ)"
        return df_transposed.reset_index()

    dict_sheets = {
        "KQ Kinh Doanh": format_dataframe(df_inc),
        "Cân Đối Kế Toán": format_dataframe(df_bal),
        "Lưu Chuyển Tiền Tệ": format_dataframe(df_cf),
    }

    print(f"    -> Đang xuất file {output_file}...")
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        has_data = False
        for sheet_name, df_out in dict_sheets.items():
            if df_out.empty: continue
            has_data = True
            df_out.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
            
        if not has_data:
            print(f"❌ Không có dữ liệu để xuất cho kỳ {period}")
            return
            
        wb = writer.book
        
        hdr_fill = PatternFill("solid", fgColor="1F4E78")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, color="1F4E78", size=14)
        item_font = Font(bold=True, size=10)
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        
        thin = Side(style="thin", color="BDD7EE")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
            period_name = "NĂM TÀI CHÍNH" if period == "year" else "CÁC QUÝ"
            ws.cell(1, 1).value = f"BÁO CÁO TÀI CHÍNH LSS ({period_name}) - CHUẨN MỰC TỐI ĐA"
            ws.cell(1, 1).font = title_font
            ws.cell(1, 1).alignment = align_center
            
            for cell in ws[2]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = align_center
                cell.border = border
                
            for row in list(ws.rows)[2:]:
                row[0].font = item_font
                row[0].alignment = align_left
                row[0].border = border
                
                for cell in row[1:]:
                    cell.border = border
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0' 
                        if cell.value < 0:
                            cell.font = Font(color="FF0000")
                    cell.alignment = align_right
                    
            ws.freeze_panes = "B3"
            
            ws.column_dimensions["A"].width = 50 
            for col_idx in range(2, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 16

    print(f"✅ Hoàn tất lưu file: {output_file}")

fetch_and_export("year", "lss_financials_all_years_max.xlsx")
fetch_and_export("quarter", "lss_financials_all_quarters_max.xlsx")
print("\n🎉 Xong toàn bộ công việc!")
