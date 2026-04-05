"""
Bảng BCTC Cộng Dồn Toàn Ngành Đường (6 công ty chính)
Cổ phiếu: SBT, QNS, LSS, SLS, KTS, FBT
Loại bỏ: NHS, SEC (theo yêu cầu)
Nguồn: sugar_group_financials.xlsx
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ─── CẤU HÌNH ────────────────────────────────────────────────────────────────
INPUT_FILE  = "sugar_group_financials.xlsx"
OUTPUT_FILE = "sugar_industry_aggregate.xlsx"
UNIT        = 1e9   # tỷ đồng
KEEP_SYMS   = ["SBT", "QNS", "LSS", "SLS", "KTS", "FBT"]

INCOME_COLS = {
    "Doanh thu thuần":                                      "Doanh thu thuần (Tỷ)",
    "Giá vốn hàng bán":                                     "Giá vốn (Tỷ)",
    "Lãi gộp":                                              "Lãi gộp (Tỷ)",
    "Chi phí bán hàng":                                     "Chi phí bán hàng (Tỷ)",
    "Chi phí quản lý DN":                                   "Chi phí QLDN (Tỷ)",
    "Lãi/Lỗ từ hoạt động kinh doanh":                     "LN HĐKD (Tỷ)",
    "LN trước thuế":                                        "LN trước thuế (Tỷ)",
    "Lợi nhuận thuần":                                      "LN thuần (Tỷ)",
    "Lợi nhuận sau thuế của Cổ đông công ty mẹ (đồng)":    "LNST cổ đông mẹ (Tỷ)",
}

BALANCE_COLS = {
    "Tài sản ngắn hạn":                    "Tổng TS ngắn hạn (Tỷ)",
    "Tài sản cố định":                     "TSCĐ (Tỷ)",
    "Tổng tài sản":                        "Tổng tài sản (Tỷ)",
    "Tổng nợ phải trả":                    "Tổng nợ (Tỷ)",
    "Vốn chủ sở hữu":                      "VCSH (Tỷ)",
}

CASHFLOW_COLS = {
    "Lưu chuyển tiền thuần trong kỳ":           "CF thuần (Tỷ)",
}

def load_sheet(sheet_name, col_map, unit=UNIT):
    df = pd.read_excel(INPUT_FILE, sheet_name=sheet_name)
    df = df[df["symbol"].isin(KEEP_SYMS)].copy()
    for c in col_map.keys():
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce") / unit
    return df

# ─── ĐỌC VÀ TỔNG HỢP ─────────────────────────────────────────────────────────
print(f"📖 Đang xử lý data cho {KEEP_SYMS}...")
df_inc = load_sheet("KQ Kinh Doanh",    INCOME_COLS)
df_bal = load_sheet("Can Doi Ke Toan",  BALANCE_COLS)
df_cf  = load_sheet("Luu Chuyen Tien Te", CASHFLOW_COLS)

def agg_df(df, col_map, group_cols):
    avail = {k: v for k, v in col_map.items() if k in df.columns}
    agg = df.groupby(group_cols)[list(avail.keys())].sum(min_count=1).reset_index()
    agg.rename(columns=avail, inplace=True)
    return agg.sort_values(group_cols)

# Tạo bảng theo Năm/Quý
agg_inc_q = agg_df(df_inc, INCOME_COLS, ["Năm", "Kỳ"])
agg_inc_y = agg_df(df_inc, INCOME_COLS, ["Năm"])
agg_bal_q = agg_df(df_bal, BALANCE_COLS, ["Năm", "Kỳ"])
agg_bal_y = agg_df(df_bal, BALANCE_COLS, ["Năm"])

# Thêm cột tính toán
for df in [agg_inc_q, agg_inc_y]:
    df["Biên lãi gộp (%)"] = (df["Lãi gộp (Tỷ)"] / df["Doanh thu thuần (Tỷ)"] * 100).round(1)
    df["Biên LNST (%)"]    = (df["LNST cổ đông mẹ (Tỷ)"] / df["Doanh thu thuần (Tỷ)"] * 100).round(1)

# Bảng so sánh LNST từng công ty theo quý
inc_pivot = df_inc.pivot_table(
    index=["Năm", "Kỳ"], columns="symbol", 
    values="Lợi nhuận sau thuế của Cổ đông công ty mẹ (đồng)", aggfunc="sum"
).reset_index()
inc_pivot[KEEP_SYMS] = inc_pivot[KEEP_SYMS].apply(pd.to_numeric) / UNIT
inc_pivot["Tổng ngành (Tỷ)"] = inc_pivot[KEEP_SYMS].sum(axis=1)

# ─── XUẤT EXCEL ──────────────────────────────────────────────────────────────
print(f"📊 Xuất file Excel: {OUTPUT_FILE}")
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    agg_inc_y.to_excel(writer, sheet_name="KQKD Ngành (Năm)", index=False, startrow=1)
    agg_inc_q.to_excel(writer, sheet_name="KQKD Ngành (Quý)", index=False, startrow=1)
    agg_bal_y.to_excel(writer, sheet_name="CĐKT Ngành (Năm)", index=False, startrow=1)
    agg_bal_q.to_excel(writer, sheet_name="CĐKT Ngành (Quý)", index=False, startrow=1)
    inc_pivot.to_excel(writer, sheet_name="LNST Từng Cty (Quý)", index=False, startrow=1)

    # Định dạng
    wb = writer.book
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    align    = Alignment(horizontal="center", vertical="center")
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Title row 1
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        ws.cell(1,1).value = f"BÁO CÁO TỔNG HỢP NGÀNH ĐƯỜNG - {sheet.upper()}"
        ws.cell(1,1).font = Font(bold=True, size=14)
        ws.cell(1,1).alignment = align
        
        # Header row 2
        for cell in ws[2]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = align
        
        # Auto column width
        for col in ws.columns:
            max_length = 0
            # Use get_column_letter(col[0].column) which is robust for MergedCells
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column_letter].width = max_length + 2

print("✅ Hoàn thành!")
