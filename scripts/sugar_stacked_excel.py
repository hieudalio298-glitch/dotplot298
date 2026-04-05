"""
Tạo file Excel dữ liệu Stacked (Cộng dồn) cho ngành đường.
Mỗi Sheet là 1 chỉ tiêu (Doanh thu, Lợi nhuận, Tài sản...).
Các cột là từng công ty + 1 cột Tổng ngành để thấy tỷ trọng đóng góp.
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ─── CẤU HÌNH ────────────────────────────────────────────────────────────────
INPUT_FILE = "sugar_group_financials.xlsx"
OUTPUT_FILE = "sugar_industry_stacked_data.xlsx"
UNIT = 1e9  # Tỷ đồng
KEEP_SYMS = ["SBT", "QNS", "LSS", "SLS", "KTS", "FBT"]

METRICS = {
    "KQ Kinh Doanh": [
        ("Doanh thu thuần", "Doanh Thu Thuần (Tỷ)"),
        ("Lãi gộp", "Lãi Gộp (Tỷ)"),
        ("Lợi nhuận sau thuế của Cổ đông công ty mẹ (đồng)", "LNST Cổ Đông Mẹ (Tỷ)"),
    ],
    "Can Doi Ke Toan": [
        ("TỔNG CỘNG TÀI SẢN (đồng)", "Tổng Tài Sản (Tỷ)"),
        ("NỢ PHẢI TRẢ (đồng)", "Tổng Nợ Phải Trả (Tỷ)"),
        ("VỐN CHỦ SỞ HỮU (đồng)", "Vốn Chủ Sở Hữu (Tỷ)"),
    ]
}

# ─── ĐỌC VÀ XỦ LÝ DỮ LIỆU ────────────────────────────────────────────────────
print(f"📖 Đọc dữ liệu từ {INPUT_FILE}...")
df_inc = pd.read_excel(INPUT_FILE, sheet_name="KQ Kinh Doanh")
df_bal = pd.read_excel(INPUT_FILE, sheet_name="Can Doi Ke Toan")

# Xử lý riêng tên cột Hàng tồn kho (có thể khác nhau)
ton_kho_col = "Hàng tồn kho, ròng (đồng)" if "Hàng tồn kho, ròng (đồng)" in df_bal.columns else "Hàng tồn kho ròng"
if ton_kho_col in df_bal.columns:
    METRICS["Can Doi Ke Toan"].append((ton_kho_col, "Hàng Tồn Kho Ròng (Tỷ)"))

def create_stacked_pivot(df, raw_col_name):
    # Lọc công ty
    df_filtered = df[df["symbol"].isin(KEEP_SYMS)].copy()
    if raw_col_name not in df_filtered.columns:
        return pd.DataFrame()
        
    df_filtered[raw_col_name] = pd.to_numeric(df_filtered[raw_col_name], errors="coerce") / UNIT
    df_filtered["Kỳ báo cáo"] = df_filtered["Năm"].astype(str) + "-Q" + df_filtered["Kỳ"].astype(str)
    
    # Pivot: Index là Kỳ, Columns là công ty
    pivot = df_filtered.pivot_table(
        index=["Năm", "Kỳ"], 
        columns="symbol", 
        values=raw_col_name, 
        aggfunc="sum"
    ).reset_index()
    
    # Sắp xếp theo Quý (tăng dần)
    pivot = pivot.sort_values(["Năm", "Kỳ"])
    pivot.insert(0, "Kỳ báo cáo", pivot["Năm"].astype(str) + "-Q" + pivot["Kỳ"].astype(str))
    pivot = pivot.drop(columns=["Năm", "Kỳ"])
    
    # Đảm bảo các cột công ty đều tồn tại dù có Null
    for sym in KEEP_SYMS:
        if sym not in pivot.columns:
            pivot[sym] = np.nan
            
    # Tính Tổng ngành
    pivot["Tổng Ngành"] = pivot[KEEP_SYMS].sum(axis=1)
    
    # Gom các cột lại theo thứ tự: Kỳ báo cáo -> Các công ty -> Tổng Ngành
    cols = ["Kỳ báo cáo"] + KEEP_SYMS + ["Tổng Ngành"]
    pivot = pivot[cols]
    
    return pivot

# ─── XUẤT EXCEL TỪNG SHEET CHỈ TIÊU ──────────────────────────────────────────
print(f"📊 Bắt đầu xuất file {OUTPUT_FILE}...")

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    sheet_data_dict = {}
    
    # Lặp qua các nhóm & khoản mục
    for sheet_source, metrics_list in METRICS.items():
        df_source = df_inc if sheet_source == "KQ Kinh Doanh" else df_bal
        for raw_col, short_name in metrics_list:
            pivot_df = create_stacked_pivot(df_source, raw_col)
            if not pivot_df.empty:
                # Cắt ngắn tên sheet (Excel giới hạn 31 ký tự)
                sheet_name = short_name.replace(" (Tỷ)", "")[:31]
                
                # Round data
                num_cols = pivot_df.select_dtypes(include=[np.number]).columns
                pivot_df[num_cols] = pivot_df[num_cols].round(1)
                
                pivot_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                sheet_data_dict[sheet_name] = short_name

    # ------------- FORMATTING -------------
    wb = writer.book
    
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill("solid", fgColor="DCE6F1")
    total_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="1E4D78")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        full_title = sheet_data_dict.get(sheet_name, sheet_name).upper() + " TỪNG CÔNG TY & TỔNG NGÀNH"
        
        # Merge Tiêu đề
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        ws.cell(1, 1).value = full_title
        ws.cell(1, 1).font = Font(bold=True, size=14, color="1F4E78")
        ws.cell(1, 1).alignment = align_center

        # Header row
        for cell in ws[2]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = align_center
            cell.border = border
            
        # Data rows
        max_col_idx = ws.max_column
        for row in list(ws.rows)[2:]:
            for idx, cell in enumerate(row):
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.0'
                    cell.alignment = align_right
                else:
                    cell.alignment = align_center
                    
                # Bôi màu riêng cho cột "Tổng Ngành" (nằm ở cuối cùng)
                if idx == max_col_idx - 1:
                    cell.fill = total_fill
                    cell.font = total_font
        
        # Freeze panes (Cố định cột đầu và 2 dòng đầu)
        ws.freeze_panes = "B3"
        
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            # Dòng 2 chứa tên header ngắn gọn
            column_letter = get_column_letter(col[1].column)
            for cell in col:
                try:
                    if cell.value:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column_letter].width = max_length + 3

print("✅ Hoàn thành!")
