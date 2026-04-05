"""
Tạo file Excel dữ liệu Stacked (Cộng dồn) cho TẤT CẢ chỉ tiêu BCTC ngành đường.
Cấu trúc mỗi Sheet (KQKD, CĐKT, LCTT):
- Dòng: Quý (Kỳ báo cáo)
- Cột lớn (Cấp 1): Tên chỉ tiêu (Doanh thu, Lợi nhuận, Phải thu, Tồn kho...)
- Cột nhỏ (Cấp 2): Từng công ty (SBT, QNS...) + Tổng Ngành
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ─── CẤU HÌNH ────────────────────────────────────────────────────────────────
INPUT_FILE = "sugar_group_financials.xlsx"
OUTPUT_FILE = "sugar_all_metrics_stacked_data.xlsx"
UNIT = 1e9  # Tỷ đồng
KEEP_SYMS = ["SBT", "QNS", "LSS", "SLS", "KTS", "FBT"]

SHEETS_TO_PROCESS = {
    "KQ Kinh Doanh": "KQKD Toan Nganh",
    "Can Doi Ke Toan": "CDKT Toan Nganh",
    "Luu Chuyen Tien Te": "LCTT Toan Nganh",
}

print(f"📖 Đọc dữ liệu từ {INPUT_FILE}...")

def process_sheet_all_metrics(sheet_name_in):
    try:
        df = pd.read_excel(INPUT_FILE, sheet_name=sheet_name_in)
    except Exception as e:
        print(f"⚠️ Không đọc được sheet {sheet_name_in}: {e}")
        return pd.DataFrame()
        
    df = df[df["symbol"].isin(KEEP_SYMS)].copy()
    if df.empty: return pd.DataFrame()
    
    # Tạo Kỳ báo cáo dạng YYYY-QX
    df["Kỳ báo cáo"] = df["Năm"].astype(str) + "-Q" + df["Kỳ"].astype(str)
    
    # Lấy các cột số (bỏ qua các cột định danh)
    skip_cols = ["symbol", "CP", "Năm", "Kỳ", "Kỳ báo cáo"]
    numeric_cols = []
    
    for c in df.columns:
        if c not in skip_cols:
            # Ép kiểu float (bỏ qua những cột text nếu có)
            df[c] = pd.to_numeric(df[c], errors="coerce")
            # Nếu cột không toàn NaN thì lấy
            if not df[c].isna().all():
                numeric_cols.append(c)
                df[c] = df[c] / UNIT  # Chuyển sang Tỷ đồng
                
    if not numeric_cols:
        return pd.DataFrame()
        
    # Melt bảng gốc: Kỳ | symbol | Chỉ_tiêu | Giá_trị
    df_melt = df.melt(id_vars=["Kỳ báo cáo", "symbol"], value_vars=numeric_cols, 
                      var_name="Chỉ tiêu", value_name="Giá trị (Tỷ)")
    
    # Pivot lại để tạo MultiIndex Columns:
    # Index: Kỳ báo cáo
    # Columns: [Chỉ tiêu, symbol]
    pivot = df_melt.pivot_table(
        index="Kỳ báo cáo", 
        columns=["Chỉ tiêu", "symbol"], 
        values="Giá trị (Tỷ)", 
        aggfunc="sum"
    )
    
    # Lấy ra danh sách các chỉ tiêu (Level 0) để thêm cột Tổng Ngành
    metrics = pivot.columns.get_level_values(0).unique()
    
    for m in metrics:
        # Nếu thiếu cty nào thì điền NaN để đủ form
        for s in KEEP_SYMS:
            if (m, s) not in pivot.columns:
                pivot[(m, s)] = np.nan
        # Tính Tổng Ngành
        # sum(axis=1) cộng gộp 6 công ty
        pivot[(m, "Tổng Ngành")] = pivot[m][KEEP_SYMS].sum(axis=1)

    # Reorder Columns: [Chỉ tiêu] -> [SBT, QNS, LSS, SLS, KTS, FBT, Tổng Ngành]
    col_order = []
    # Sắp xếp các chỉ tiêu theo thứ tự xuất hiện ban đầu trong file gốc cho hợp logic BCTC
    ordered_metrics = [m for m in numeric_cols if m in metrics]
    
    for m in ordered_metrics:
        for s in KEEP_SYMS + ["Tổng Ngành"]:
            col_order.append((m, s))
            
    pivot = pivot[col_order].sort_index()
    # Chỉ lấy 20 quý gần nhất nếu quá nhiều
    # pivot = pivot.tail(20)  # Có thể bật lên nếu user muốn xem ngắn
    
    return pivot


# ─── XUẤT EXCEL ──────────────────────────────────────────────────────────────
print(f"📊 Bắt đầu tạo ma trận data cho toàn bộ chỉ tiêu...")

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    for sheet_in, sheet_out in SHEETS_TO_PROCESS.items():
        print(f"   - Xử lý mảng: {sheet_in}")
        pivot_df = process_sheet_all_metrics(sheet_in)
        if pivot_df.empty: 
            continue
            
        # Làm tròn
        pivot_df = pivot_df.round(2)
        
        # Ghi vào Excel
        pivot_df.to_excel(writer, sheet_name=sheet_out)
        
    wb = writer.book
    
    # Formatting
    hdr1_fill = PatternFill("solid", fgColor="0B1E33")
    hdr2_fill = PatternFill("solid", fgColor="1F4E78")
    total_fill = PatternFill("solid", fgColor="DCE6F1")
    
    font_bold_w = Font(bold=True, color="FFFFFF")
    font_bold_b = Font(bold=True, color="000000")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin = Side(style="thin", color="1E4D78")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Sửa tên index (A1, A2) để thân thiện hơn
        ws["A1"] = "Khoản Mục BCTC"
        ws["A2"] = "Kỳ Báo Cáo"
        
        # Format Header Row 1 (Tên Chỉ Tiêu)
        for cell in ws[1]:
            cell.fill = hdr1_fill
            cell.font = font_bold_w
            cell.alignment = align_center
            cell.border = border
            
        # Format Header Row 2 (Tên Công Ty)
        for cell in ws[2]:
            cell.fill = hdr2_fill
            cell.font = font_bold_w
            cell.alignment = align_center
            cell.border = border
            
        # Format Cột A (Kỳ báo cáo)
        for row in list(ws.rows)[2:]:
            row[0].font = font_bold_b
            row[0].alignment = align_center
            row[0].border = border
            
        # Dữ liệu & Cột Tổng
        cols_count = ws.max_column
        for row_idx in range(3, ws.max_row + 1):
            for col_idx in range(2, cols_count + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.0'
                    cell.alignment = align_right
                else:
                    cell.alignment = align_center
                    
                # Bôi màu cột Tổng Ngành
                # Dựa vào tiêu đề cột row 2
                col_name = ws.cell(row=2, column=col_idx).value
                if col_name == "Tổng Ngành":
                    cell.fill = total_fill
                    cell.font = font_bold_b
                    
        ws.freeze_panes = "B3"
        
        # Auto-fit (căn chỉnh chiều rộng)
        # Chỉ tiêu gộp nhiều ô (merged_cells) => set độ rộng vừa phải
        ws.column_dimensions["A"].width = 15
        for col_idx in range(2, cols_count + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 12

print(f"✅ Đã lưu file: {OUTPUT_FILE}")
