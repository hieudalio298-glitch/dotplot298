"""
Script lấy BCTC của riêng mã LSS - Tất cả các năm và các Quý có thể lấy được.
Sử dụng nguồn KBS.
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from vnstock import Vnstock

SYMBOL = "LSS"
SOURCE = "KBS"

print("=" * 70)
print(f"  TẢI BCTC {SYMBOL} - TOÀN BỘ LỊCH SỬ NĂM/QUÝ (Nguồn {SOURCE})")
print("=" * 70)

fin = Vnstock().stock(symbol=SYMBOL, source=SOURCE).finance

def fetch_and_export(period, output_file):
    print(f"\n[+] Đang tải dữ liệu {period.upper()}...")
    df_inc = fin.income_statement(period=period)
    df_bal = fin.balance_sheet(period=period)
    df_cf = fin.cash_flow(period=period)
    
    def format_dataframe(df):
        if df is None or df.empty:
            return pd.DataFrame()
            
        df = df.copy()
        
        if "item_id" in df.columns:
            df = df.drop(columns=["item_id"])
            
        if "item" in df.columns:
            df = df.rename(columns={"item": "Chỉ tiêu (VNĐ)"})
            
        # Tìm tất cả các cột chứa số (Năm) hoặc có ký tự Q (Quý)
        time_cols = [c for c in df.columns if str(c).isdigit() or ("-" in str(c) and "Q" in str(c))]
        time_cols_sorted = sorted(time_cols, reverse=True)
        
        final_cols = ["Chỉ tiêu (VNĐ)"] + time_cols_sorted
        df = df[final_cols]
        
        return df

    dict_sheets = {
        "KQ Kinh Doanh": format_dataframe(df_inc),
        "Cân Đối Kế Toán": format_dataframe(df_bal),
        "Lưu Chuyển Tiền Tệ": format_dataframe(df_cf),
    }

    print(f"    -> Đang xuất file {output_file}...")
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        has_data = False
        for sheet_name, df_out in dict_sheets.items():
            if df_out.empty: continue
            has_data = True
            df_out.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
            
        if not has_data:
            print(f"❌ Không có dữ liệu để xuất cho kỳ {period}")
            return
            
        wb = writer.book
        
        hdr_fill = PatternFill("solid", fgColor="1F4E78")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, color="1F4E78", size=14)
        item_font = Font(bold=True, size=10)
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        
        thin = Side(style="thin", color="BDD7EE")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
            period_name = "NĂM TÀI CHÍNH" if period == "year" else "CÁC QUÝ"
            ws.cell(1, 1).value = f"BÁO CÁO TÀI CHÍNH LSS ({period_name}) - CHUẨN VAS"
            ws.cell(1, 1).font = title_font
            ws.cell(1, 1).alignment = align_center
            
            for cell in ws[2]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = align_center
                cell.border = border
                
            for row in list(ws.rows)[2:]:
                row[0].font = item_font
                row[0].alignment = align_left
                row[0].border = border
                
                for cell in row[1:]:
                    cell.border = border
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0' 
                        if cell.value < 0:
                            cell.font = Font(color="FF0000")
                    cell.alignment = align_right
                    
            ws.freeze_panes = "B3"
            
            ws.column_dimensions["A"].width = 50 
            for col_idx in range(2, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 16

    print(f"✅ Hoàn tất lưu file: {output_file}")

# Tải theo năm
fetch_and_export("year", "lss_financials_all_years_vas_v2.xlsx")

# Tải theo quý
fetch_and_export("quarter", "lss_financials_all_quarters_vas.xlsx")

print("\n🎉 Xong toàn bộ công việc!")
