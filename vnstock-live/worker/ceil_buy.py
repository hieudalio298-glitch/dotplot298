"""
Tính dư mua trần (Ceiling Buying Excess) phiên hôm nay.
Xuất file Excel với định dạng đẹp.
Dư mua trần = Tổng KL chào mua ở giá trần - Tổng KL chào bán ở giá trần
cho tất cả các mã đang match giá trần hoặc có KL mua tại trần.
"""

import re
import sys
from datetime import date
from pathlib import Path
import pandas as pd
import vnstock

from supabase import create_client

SUPABASE_URL = "https://utqmpdmbkubhzuccqeyf.supabase.co"
SUPABASE_KEY = "sb_publishable_udSp3L071S-ShERkLoCcjw_HO4JD4is"


def fetch_all_symbols():
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    all_data = []
    page = 0
    while True:
        r = sb.from_("stocks").select("symbol").range(page, page + 999).execute()
        if not r.data:
            break
        all_data.extend(r.data)
        if len(r.data) < 1000:
            break
        page += 1000
    return [d["symbol"] for d in all_data if re.match(r'^[A-Z]{3}$', d["symbol"])]


def main():
    symbols = fetch_all_symbols()
    print(f"Total symbols: {len(symbols)}")

    all_boards = []
    for i in range(0, len(symbols), 50):
        batch = symbols[i:i+50]
        try:
            board = vnstock.Trading(source="VCI").price_board(batch)
            if board is not None and not board.empty:
                if isinstance(board.columns, pd.MultiIndex):
                    board.columns = ["_".join(str(c) for c in col).strip() for col in board.columns.values]
                all_boards.append(board)
        except Exception as e:
            pass

    df = pd.concat(all_boards, ignore_index=True)
    print(f"Total rows fetched: {len(df)}")

    # Parse numeric columns
    for c in df.columns:
        if "price" in c or "volume" in c or "count" in c or "floor" in c or "ceil" in c or "ref" in c:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    results = []
    total_ceil_buy = 0
    total_ceil_sell = 0

    for _, row in df.iterrows():
        sym = str(row.get("listing_symbol", ""))
        ceil_price = float(row.get("listing_ceiling", 0))
        match_price = float(row.get("match_match_price", 0))
        ref_price = float(row.get("listing_ref_price", 0))
        total_vol = float(row.get("match_accumulated_volume", 0))

        if ceil_price <= 0 or match_price <= 0:
            continue

        # Check buy side at ceiling price
        buy_vol_at_ceil = 0
        for level in [1, 2, 3]:
            bid_price = float(row.get(f"bid_ask_bid_{level}_price", 0))
            bid_vol = float(row.get(f"bid_ask_bid_{level}_volume", 0))
            if abs(bid_price - ceil_price) < 0.01:
                buy_vol_at_ceil += bid_vol

        # Check sell side at ceiling price
        sell_vol_at_ceil = 0
        for level in [1, 2, 3]:
            ask_price = float(row.get(f"bid_ask_ask_{level}_price", 0))
            ask_vol = float(row.get(f"bid_ask_ask_{level}_volume", 0))
            if abs(ask_price - ceil_price) < 0.01:
                sell_vol_at_ceil += ask_vol

        # Stock at ceiling or has buy orders at ceiling
        at_ceil = abs(match_price - ceil_price) < 0.01
        has_ceil_activity = buy_vol_at_ceil > 0 or sell_vol_at_ceil > 0 or at_ceil

        if has_ceil_activity:
            excess = buy_vol_at_ceil - sell_vol_at_ceil  # Dương = dư mua, Âm = dư bán
            change_pct = ((match_price - ref_price) / ref_price * 100) if ref_price > 0 else 0
            results.append({
                "symbol": sym,
                "ceil_price": ceil_price,
                "match_price": match_price,
                "change_pct": change_pct,
                "at_ceil": at_ceil,
                "buy_at_ceil": buy_vol_at_ceil,
                "sell_at_ceil": sell_vol_at_ceil,
                "excess": excess,
                "total_vol": total_vol,
                "buy_value_ty": round(buy_vol_at_ceil * ceil_price / 1e9, 2),
                "excess_value_ty": round(excess * ceil_price / 1e9, 2),
            })
            total_ceil_buy += buy_vol_at_ceil
            total_ceil_sell += sell_vol_at_ceil

    # Sort by excess buy volume (largest first)
    results.sort(key=lambda x: x["excess"], reverse=True)

    total_buy_value_ty = sum(r["buy_value_ty"] for r in results)
    total_excess_value_ty = sum(r["excess_value_ty"] for r in results)

    print(f"\n{'='*80}")
    print(f"DU MUA TRAN PHIEN HOM NAY")
    print(f"{'='*80}")
    print(f"\nTong ma co hoat dong tai gia tran: {len(results)}")
    print(f"Tong KL mua tran: {total_ceil_buy:,.0f}")
    print(f"Tong KL ban tran: {total_ceil_sell:,.0f}")
    print(f"DU MUA TRAN TOAN THI TRUONG: {total_ceil_buy - total_ceil_sell:,.0f}")
    print(f"GIA TRI DU MUA TRAN: {total_excess_value_ty:,.1f} ty dong")
    print(f"{'='*80}")

    # Top excess buyers
    top_buy = [r for r in results if r["excess"] > 0]
    top_buy.sort(key=lambda x: x["excess"], reverse=True)

    if top_buy:
        print(f"\nTOP 20 MA DU MUA TRAN LON NHAT:")
        print(f"{'Ma':<6} {'Gia Tran':>10} {'Gia Khop':>10} {'%':>7} {'Mua Tran':>12} {'Du Mua':>12} {'GT Du Mua(ty)':>14} {'Tong KL':>12}")
        print(f"{'-'*95}")
        for r in top_buy[:20]:
            flag = " <<< TRAN" if r["at_ceil"] else ""
            print(
                f"{r['symbol']:<6} {r['ceil_price']/1000:>9.1f}K {r['match_price']/1000:>9.1f}K "
                f"{r['change_pct']:>+6.1f}% {r['buy_at_ceil']:>11,.0f} "
                f"{r['excess']:>11,.0f} {r['excess_value_ty']:>13.1f} {r['total_vol']:>11,.0f}{flag}"
            )

    # Stocks AT ceiling price
    at_ceil_list = [r for r in results if r["at_ceil"]]
    if at_ceil_list:
        print(f"\n{'='*80}")
        print(f"CO PHIEU DANG KHOP GIA TRAN ({len(at_ceil_list)} ma):")
        print(f"{'Ma':<6} {'Gia Tran':>10} {'Mua Tran':>12} {'Ban Tran':>12} {'Du Mua':>12} {'Tong KL':>12}")
        print(f"{'-'*75}")
        at_ceil_list.sort(key=lambda x: x["excess"], reverse=True)
        for r in at_ceil_list:
            print(
                f"{r['symbol']:<6} {r['ceil_price']/1000:>9.1f}K "
                f"{r['buy_at_ceil']:>11,.0f} {r['sell_at_ceil']:>11,.0f} "
                f"{r['excess']:>11,.0f} {r['total_vol']:>11,.0f}"
            )

    # ── EXCEL EXPORT ──────────────────────────────────────────
    today_str = date.today().strftime("%Y-%m-%d")
    out_file = f"du_mua_tran_{today_str}.xlsx"

    all_rows = [{
        "Ma": r["symbol"],
        "Gia Tran (K)": round(r["ceil_price"] / 1000, 1),
        "Gia Khop (K)": round(r["match_price"] / 1000, 1),
        "Thay doi (%)": round(r["change_pct"], 2),
        "Tai Tran": "CO" if r["at_ceil"] else "",
        "KL Mua Tran": int(r["buy_at_ceil"]),
        "KL Ban Tran": int(r["sell_at_ceil"]),
        "Du Mua Tran (cp)": int(r["excess"]),
        "GT Du Mua Tran (ty)": r["excess_value_ty"],
        "Tong KL Ngay": int(r["total_vol"]),
    } for r in results]

    df_out = pd.DataFrame(all_rows)
    df_out = df_out.sort_values("Du Mua Tran (cp)", ascending=False)

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        # Sheet 1: Summary
        summary = pd.DataFrame([
            {"Chi tieu": "Ngay giao dich", "Gia tri": today_str},
            {"Chi tieu": "Tong ma co hoat dong tai tran", "Gia tri": len(results)},
            {"Chi tieu": "So ma dang khop tran", "Gia tri": len(at_ceil_list)},
            {"Chi tieu": "Tong KL mua tran (cp)", "Gia tri": int(total_ceil_buy)},
            {"Chi tieu": "Tong KL ban tran (cp)", "Gia tri": int(total_ceil_sell)},
            {"Chi tieu": "Du mua tran (cp)", "Gia tri": int(total_ceil_buy - total_ceil_sell)},
            {"Chi tieu": "Gia tri mua tran (ty dong)", "Gia tri": round(total_buy_value_ty, 1)},
            {"Chi tieu": "GIA TRI DU MUA TRAN (ty dong)", "Gia tri": round(total_excess_value_ty, 1)},
        ])
        summary.to_excel(writer, sheet_name="Tong hop", index=False)

        # Sheet 2: All data
        df_out.to_excel(writer, sheet_name="Chi tiet", index=False)

        # Sheet 3: At ceiling only
        df_ceil = df_out[df_out["Tai Tran"] == "CO"].copy()
        if not df_ceil.empty:
            df_ceil.to_excel(writer, sheet_name="Dang khop tran", index=False)

        # Format
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            # Header bold
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            # Auto column width
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)
            # Highlight "at ceiling" rows in Chi tiet sheet with GREEN
            if sheet_name == "Chi tiet":
                green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                for row in ws.iter_rows(min_row=2):
                    if row[4].value == "CO":  # Tai Tran column
                        for cell in row:
                            cell.fill = green_fill

    print(f"\n>> Excel da luu: {out_file}")


if __name__ == "__main__":
    main()
